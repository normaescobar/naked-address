from openpyxl.styles import colors, PatternFill, Font, Color
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from decimal import Decimal
import requests
import shutil
import json
import time
import os
import re

INPUT_PATH = 'input/'
OUTPUT_PATH = 'output/'


HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,
    'referer':'https://www.google.com/'
}

ASSET_TX_URLS = {
    'BTC' : 'https://www.blockchain.com/explorer/transactions/btc/',
    'BCH' : 'https://www.blockchain.com/explorer/transactions/bch/',
    'DASH' : 'https://explorer.dash.org/insight/tx/',
    'DOGE' : 'https://dogechain.info/tx/',
    'LTC' : 'https://litecoinblockexplorer.net/tx/',
    'ZEC' : 'https://explorer.zcha.in/transactions/',
    'ALGO' : 'https://algoexplorer.io/tx/',
    'XRP' : 'https://xrpscan.com/tx/',
    'XLM' : 'https://stellarchain.io/transactions/',
    'ETC' : 'https://etcblockexplorer.com/tx/'
}

def set_verified_tx(work_sheet, row, column, value):
    work_sheet[row][column].value = f'=HYPERLINK("{value}", "Verified")'
    work_sheet[row][column].fill = PatternFill(start_color='C3ECCB', end_color='C3ECCB', fill_type = 'solid')
    work_sheet[row][column].font = Font(color='006100')

def set_worksheet(output, asset):
    ws_count = len(output.sheetnames)
    work_sheet = output.active
    if ws_count == 1 and work_sheet.title == 'Sheet':
        work_sheet.title = asset
    else:
        output.create_sheet(asset)
        output.active = ws_count
        work_sheet = output.active
        
    return work_sheet
    
def get_blockexplorer_pagecount(api_link, address, tx_keys):
    link = f'{api_link}address/{address}'
    is_valid = requests.get(link, headers=HEADERS)
    
    if is_valid.status_code != 200:
        return 0
        
    content = json.loads(is_valid.content)
    
    tx_count = 0
    
    for tx_key in tx_keys:
        tx_count += content[tx_key]
        
    if tx_count == 0:
        return tx_count
    
    return content['totalPages']
        
def process_stellar(asset, asset_symbol, asset_decimal, address, start_date, end_date, output):
    api_link = f'https://horizon.stellar.org/accounts/{address}/transactions?limit=200'

    txns = requests.get(api_link, headers=HEADERS)
    content = json.loads(txns.content)

    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Blockchain URL'])        
    
    rows = content['_embedded']['records']
    cursor = re.search(r'cursor=(\d+)', content['_links']['next']['href']).group(1)
    
    while True:
        for row in rows:
            timestamp = int(datetime.strptime(row['created_at'].split('T')[0], "%Y-%m-%d").timestamp())
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            tx_link = f'https://horizon.stellar.org/transactions/{row["id"]}/operations'
            tx_details = requests.get(tx_link, headers=HEADERS)
            tx_content = json.loads(tx_details.content)
            data = tx_content['_embedded']['records'][0]
            if data['type'] != 'payment':
                continue 
            sent = Decimal(data['amount']) if data['from'] == address else 0
            received = Decimal(data['amount']) if data['to'] == address else 0
            work_sheet.append([row['created_at'], row['ledger'], row['id'], sent, received, asset_symbol, ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{ASSET_TX_URLS[asset_symbol]}{row["id"]}')        
        
        txns = requests.get(f'{api_link}&cursor={cursor}', headers=HEADERS)
        content = json.loads(txns.content)
        rows = content['_embedded']['records']
        cursor = re.search(r'cursor=(\d+)', content['_links']['next']['href']).group(1)
        if cursor == '':
            break
            
    
def process_ripple(asset, asset_symbol, asset_decimal, address, start_date, end_date, output):
    api_link = f'https://api.xrpscan.com/api/v1/account/{address}/transactions'

    txns = requests.get(api_link, headers=HEADERS)
    
    if txns.status_code != 200:
        return
        
    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Blockchain URL'])        
        
    content = json.loads(txns.content)
    
    rows = content['transactions']
    marker = content['marker'] if 'marker' in content else None
    
    while True:
        for row in rows:  
            timestamp = int(datetime.strptime(row['date'].split('T')[0], "%Y-%m-%d").timestamp())
            if timestamp < start_date:
                break
            if timestamp > end_date:
                continue
                
            if row['TransactionType'] != 'Payment':
                continue
            value = int(row['Amount']['value'])
            currency = row['Amount']['currency']
            sent = value/asset_decimal if row['Destination'] != address else 0
            received = value/asset_decimal if row['Destination'] == address else 0
            work_sheet.append([row['date'], row['ledger_index'], row['hash'], sent, received, currency, ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{ASSET_TX_URLS[asset_symbol]}{row["hash"]}')
            
        txns = requests.get(f'{api_link}?marker={marker}', headers=HEADERS)
        if txns.status_code != 200:
            return        
        content = json.loads(txns.content)
        rows = content['transactions']
        marker = content['marker'] if 'marker' in content else None
        if marker is None:
            break


def process_algorand(asset, asset_symbol, asset_decimal, address, start_date, end_date, output):
    before_time = datetime.fromtimestamp(end_date + 86400 ).strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    after_time = datetime.fromtimestamp(start_date - 86400).strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    api_link = f'https://algoindexer.algoexplorerapi.io/v2/transactions?address={address}&before-time={before_time}&after-time={after_time}'
    txns = requests.get(api_link)
   
    content = json.loads(txns.content)
   
    if 'message' in content:
        return
   
    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Blockchain URL'])
   
    rows = content['transactions']
    next_token = content['next-token'] if 'next-token' in content else None
    
    while True:
        for row in rows:
            timestamp = int(row['round-time'])
            date = str(datetime.fromtimestamp(timestamp))
            value = int(row['payment-transaction']['amount'])
            sent = value/asset_decimal if row['payment-transaction']['receiver'].lower() != address.lower() else 0
            received = value/asset_decimal if row['payment-transaction']['receiver'].lower() == address.lower() else 0
           
            work_sheet.append([date, row['confirmed-round'], row['id'], sent, received, asset_symbol, ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{ASSET_TX_URLS[asset_symbol]}{row["id"]}')
        
        txns = requests.get(f'{api_link}&next={next_token}')
        content = json.loads(txns.content)
        rows = content['transactions']
        next_token = content['next-token'] if 'next-token' in content else None
        if len(rows) < 1:
            break   


def get_eth_block(api_link):
    tx = requests.get(api_link)
    content = json.loads(tx.content)
    return int(content['result'][0]['blockNumber'])

def process_ethereum(asset, asset_symbol, asset_decimal, address, start_date, end_date, output):
    api_key = 'XQBQNA7X7Z8BU41N9DZ5WUGT5E7YSFF15M'
    api_link = f'https://api.etherscan.io/api?module=account&address={address}&apikey={api_key}'
    
    eth_tx_url = 'https://etherscan.io/tx/'
    address = address.lower()
    
    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Blockchain URL']) 
    
    start_block = get_eth_block(f'{api_link}&action=txlist&startblock=0&endblock=99999999&sort=asc&offset=1&page=1')
    end_block = get_eth_block(f'{api_link}&action=txlist&startblock=0&endblock=99999999&sort=desc&offset=1&page=1')
    
    block_increment = 100000
    while start_block <= end_block:
        print(f'Working on blocks {start_block}-{start_block + block_increment}')
        api_txlist = f'{api_link}&action=txlist&startblock={start_block}&endblock={start_block + block_increment}&sort=asc'
        api_tokentx = f'{api_link}&action=tokentx&startblock={start_block}&endblock={start_block + block_increment}&sort=asc'

        tx_list = requests.get(api_txlist)
        content = json.loads(tx_list.content)
        rows = content['result']
        for row in rows:
            timestamp = int(row['timeStamp'])
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            value = Decimal(row['value'])
            if value > 0:
                value /= Decimal(asset_decimal)
                sent = 0
                received = 0
                if address == row['from']:
                    sent = value
                elif address == row['to']:
                    received = value
                    
                work_sheet.append([str(datetime.fromtimestamp(timestamp)), row['blockNumber'], row['hash'], sent, received, asset_symbol, ''])
                set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{eth_tx_url}{row["hash"]}')

        token_tx = requests.get(api_tokentx)
        content = json.loads(token_tx.content)
        rows = content['result']
        for row in rows:
            token_symbol = row['tokenSymbol']
        
            timestamp = int(row['timeStamp'])
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            value = Decimal(row['value']) / Decimal(10 ** int(row['tokenDecimal']))
            sent = 0
            received = 0
            if address == row['from']:
                sent = value
            elif address == row['to']:
                received = value
                    
            work_sheet.append([str(datetime.fromtimestamp(timestamp)), row['blockNumber'], row['hash'], sent, received, token_symbol, ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{eth_tx_url}{row["hash"]}')
        
        start_block += block_increment + 1
        
                
def process_blockchair_asset(asset, asset_symbol, asset_decimal, api_keyword, address, start_date, end_date, output):
    start_date = datetime.fromtimestamp(start_date).strftime('%Y-%m-%d')
    end_date = datetime.fromtimestamp(end_date).strftime('%Y-%m-%d')
    txns = requests.get(f'https://api.blockchair.com/{api_keyword}/dashboards/address/{address}?transaction_details=true&q=time({start_date}..{end_date})')
   
    content = json.loads(txns.content)
   
    if content['data'] is None or content['data'][address]['address']['type'] is None:
        return
       
    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Blockchain URL'])
   
    rows = content['data'][address]['transactions']
   
    for row in rows:
        value = int(row['balance_change'])
        sent = value*-1/asset_decimal if value < 0 else 0
        received = value/asset_decimal if value > 0 else 0
        
       
        work_sheet.append([row['time'], row['block_id'], row['hash'], sent, received, asset_symbol, ''])
        set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{ASSET_TX_URLS[asset_symbol]}{row["hash"]}')    


def process_address(asset, address, start_date, end_date, output):
    if asset is None:
        return
        
    if asset.lower() == 'bitcoin':
        process_blockchair_asset(asset, 'BTC', 1e8, 'bitcoin', address, start_date, end_date, output)
    elif asset.lower() == 'bitcoin cash':
        process_blockchair_asset(asset, 'BCH', 1e8, 'bitcoin-cash', address, start_date, end_date, output)
    elif asset.lower() == 'dogechain':
        process_blockchair_asset(asset, 'DOGE', 1e8, 'dogecoin', address, start_date, end_date, output)
    elif asset.lower() == 'dash':
        process_blockchair_asset(asset, 'DASH', 1e8, 'dash', address, start_date, end_date, output)
    elif asset.lower() == 'litecoin':
        process_blockchair_asset(asset, 'LTC', 1e8, 'litecoin', address, start_date, end_date, output)
    elif asset.lower() == 'zcash':
        process_blockchair_asset(asset, 'ZEC', 1e8, 'zcash', address, start_date, end_date, output)
    elif asset.lower() == 'algorand':
        process_algorand(asset, 'ALGO', 1e6, address, start_date, end_date, output)
    # elif asset.lower() == 'ripple':
        # process_ripple(asset, 'XRP', 1e6, address, start_date, end_date, output)
    #elif asset.lower() == 'stellar':
    #    process_stellar(asset, 'XLM', 1, address, start_date, end_date, output)
    if asset.lower() == 'ethereum':
       process_ethereum(asset, 'ETH', 1e18, address, start_date, end_date, output)
    


def process_input(filename, start_date, end_date, output):
    file = f'{INPUT_PATH}{filename}'
    work_book = load_workbook(file, data_only=True)
    
    for worksheet_idx in range(len(work_book.worksheets)):
        asset_col = None
        addr_col = None
        work_book.active = worksheet_idx
        work_sheet = work_book.active
        print(f'Working with sheet {work_sheet.title}')
       
        for column_idx, column_name in enumerate(work_sheet[1]):
            if column_name.value == 'Asset':
                asset_col = column_idx
            elif column_name.value == 'Address':
                addr_col = column_idx


        row_count = work_sheet.max_row
       
        for row_idx in range(2, row_count + 1):
            print(f'Working with row #{row_idx}/{row_count}')
            if asset_col is not None and addr_col is not None:
                asset = work_sheet[row_idx][asset_col].value
                address = work_sheet[row_idx][addr_col].value
            else:
                print(f'asset column or address column not found')
                break
                
            process_address(asset, address, start_date, end_date, output)
            
                
            
                 
    print('Processing Done.')
    
if __name__ == '__main__':
    start_date = int(datetime.strptime(input('Start Date (m/d/yyyy): '), "%m/%d/%Y").timestamp())
    end_date = int(datetime.strptime(input('End Date (m/d/yyyy): '), "%m/%d/%Y").timestamp())
    
    _, _, files = next(os.walk(INPUT_PATH))

    output = Workbook()
    
    for file in files:
        process_input(file, start_date, end_date, output)
        
    output.save(f'{OUTPUT_PATH}output.xlsx')
